import subprocess
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from typing import List, Dict

def get_git_urls() -> List[str]:
    """Get list of git URLs from user input."""
    urls = []
    while True:
        url = input("Enter git URL (one at a time, empty line to finish): ").strip()
        if not url:
            break
        urls.append(url)
    return urls

def run_cli_command(git_url: str) -> Dict:
    """Run CLI command for given git URL and return JSON result."""
    command = f"veracode scan --source {git_url} --type repo --format json"
    try:
        result = subprocess.run(command, shell=True, capture_output=True, text=True)
        return json.loads(result.stdout)
    except Exception as e:
        print(f"Error running command for {git_url}: {e}")
        return {"vulnerabilities": [], "secrets": [], "configs": []}
    
def format_fixes(fix_data: Dict) -> str:
    """Format available fixes step-by-step."""
    if "available" not in fix_data:
        return "N/A"
    
    available_fixes = fix_data.get("available", [])
    fix_strings = []
    
    for available_fix in available_fixes:
        version = available_fix.get('version', 'N/A')
        date = available_fix.get('date', 'N/A')
        fix_strings.append(f"{version} ({date})")
    
    return ", ".join(fix_strings)

def safe_get(dictionary, key, default="N/A") -> str:
        return f"{key} -> {str(dictionary)}" if not isinstance(dictionary, dict) else dictionary.get(key, default)

def process_results(git_urls: List[str]) -> Dict[str, List[Dict]]:
    """Process results from CLI for all URLs."""
    parsed_results = {"raw_json": [], "vulnerabilities": [], "secrets": [], "configs": []}
    
    for url in git_urls:
        print(f"Processing {url}...")
        cli_results = run_cli_command(url)

        parsed_results["raw_json"].append({
            "git_url": url,
            "raw_json": json.dumps(cli_results)
        })

        vulnerabilities = cli_results.get("vulnerabilities", {})
        if vulnerabilities:
            for vuln in vulnerabilities.get("matches", []):
                parsed_results["vulnerabilities"].append({
                    "git_url": url,
                    "id": safe_get(vuln["vulnerability"], "id"),
                    "description": safe_get(vuln["vulnerability"], "description"),
                    "artifact_name": safe_get(vuln["artifact"], "name") if "artifact" in vuln else "N/A",
                    "artifact_version": safe_get(vuln["artifact"], "version") if "artifact" in vuln else "N/A",
                    "artifact_type": safe_get(vuln["artifact"], "type") if "artifact" in vuln else "N/A",
                    "risk": safe_get(vuln["vulnerability"], "risk"),
                    "severity": safe_get(vuln["vulnerability"], "severity"),
                    "CVEs": (", ".join([f"{cwe['cve']} ({cwe['cwe']})" for cwe in safe_get(vuln["vulnerability"], "cwes", [])])) if safe_get(vuln["vulnerability"], "cwes") else "N/A",
                    "fixes": format_fixes(safe_get(vuln["vulnerability"], "fix")),
                    "raw_json": str(vuln)
                })
        
        secrets = safe_get(cli_results, "secrets", [])
        if isinstance(secrets, list):
            for secret in secrets:
                parsed_results["secrets"].append({
                    "git_url": url,
                    "category": safe_get(secret, "Category"),
                    "severity": safe_get(secret, "Severity"),
                    "start_line": safe_get(secret, "StartLine"),
                    "end_line": safe_get(secret, "EndLine"),
                    "match": safe_get(secret, "Match"),
                    "file_path": safe_get(secret, "Target"),
                    "code": "\n".join([line["Content"] for line in secret["Code"]["Lines"]]) if "Code" in secret else "N/A",
                    "raw_json": str(secret)
                })
        else:
           results = safe_get(secrets, "Results", [])
           for result in results:
                inner_secrets = safe_get(result, "Secrets", [])
                for secret in inner_secrets:
                    parsed_results["secrets"].append({
                        "git_url": url,
                        "category": safe_get(secret, "Category"),
                        "severity": safe_get(secret, "Severity"),
                        "start_line": safe_get(secret, "StartLine"),
                        "end_line": safe_get(secret, "EndLine"),
                        "match": safe_get(secret, "Match"),
                        "file_path": safe_get(result, "Target"),
                        "code": "\n".join([line["Content"] for line in secret["Code"]["Lines"]]) if "Code" in secret else "N/A",
                        "raw_json": str(secret)
                    })

        
        configs = safe_get(cli_results, "configs", [])
        if isinstance(secrets, list):
            for config in configs:
                parsed_results["configs"].append({
                    "git_url": url,
                    "target_type": "N/A",
                    "avdid": safe_get(config, "AVDID", "N/A"),
                    "id": "N/A",
                    "severity": safe_get(config, "Severity", "N/A"),
                    "status": safe_get(config, "Status", "N/A"),
                    "message": safe_get(config, "Message", "N/A"),
                    "namespace": safe_get(config, "Namespace", "N/A"),
                    "primary_url": safe_get(config, "PrimaryURL", "N/A"),
                    "query": safe_get(config, "Query", "N/A"),
                    "target": safe_get(config, "Target", "N/A"),
                    "title": safe_get(config, "Title", "N/A"),
                    "type": safe_get(config, "Type", "N/A"),                
                    "description": safe_get(config, "Description", "N/A"),
                    "resolution": safe_get(config, "Resolution", "N/A"),
                    "raw_json": str(config)
                })            
        else:
            results = safe_get(configs, "Results", [])
            if results:
                for result in results:
                    for misconfiguration in safe_get(result, "Misconfigurations", []):
                        parsed_results["configs"].append({
                            "git_url": url,
                            "target_type": safe_get(result, "Type", "N/A"),
                            "avdid": safe_get(misconfiguration, "AVDID", "N/A"),
                            "id": safe_get(misconfiguration, "ID", "N/A"),
                            "severity": safe_get(misconfiguration, "Severity", "N/A"),
                            "status": safe_get(misconfiguration, "Status", "N/A"),
                            "message": safe_get(misconfiguration, "Message", "N/A"),
                            "namespace": safe_get(misconfiguration, "Namespace", "N/A"),
                            "primary_url": safe_get(misconfiguration, "PrimaryURL", "N/A"),
                            "query": safe_get(misconfiguration, "Query", "N/A"),
                            "target": safe_get(result, "Target", "N/A"),
                            "title": safe_get(misconfiguration, "Title", "N/A"),
                            "type": safe_get(misconfiguration, "Type", "N/A"),
                            "description": safe_get(misconfiguration, "Description", "N/A"),
                            "resolution": safe_get(misconfiguration, "Resolution", "N/A"),
                            "raw_json": str(misconfiguration)   
                        })
            
    
    return parsed_results

def calculate_cell_length(value: str) -> int:
    """Calculate appropriate cell length for Excel column width."""
    lines = value.splitlines()
    max_length = max(len(line) for line in lines)
    return max_length

def save_to_excel(results: Dict[str, List[Dict]], output_file: str = "combined_iac_report.xlsx"):
    """Save results to Excel with separate tabs using openpyxl."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    text_wrap = Alignment(wrap_text=True)
    
    for sheet_name, data in results.items():
        worksheet = workbook.create_sheet(title=sheet_name.capitalize())
        
        if data:
            # Write headers
            headers = list(data[0].keys())
            worksheet.append([header.replace("_", " ").capitalize() for header in headers])            
            
            # Write data rows
            for row_data in data:
                row_values = [row_data.get(header, "") for header in headers]
                worksheet.append(row_values)
            
            # Apply text wrapping
            dimensions = {}
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if cell.value:
                        dimensions[cell.column_letter] = max(dimensions.get(cell.column_letter, 0), calculate_cell_length(str(cell.value)))
                    if cell.col_idx < worksheet.max_column:
                        cell.alignment = text_wrap
            for col, width in dimensions.items():
                worksheet.column_dimensions[col].width = width + 2

    workbook.save(output_file)
    print(f"Report saved to {output_file}")

def main():
    """Main execution flow."""
    git_urls = get_git_urls()
    file_name = input("Enter the name of the xlsx file to save (defaults to combined_iac_report.xlsx): ")
    if not file_name or file_name.strip() == "":
        file_name = "combined_iac_report.xlsx"
    elif not file_name.endswith(".xlsx"):
        file_name += ".xlsx"

    if not git_urls:
        print("No URLs provided.")
        return

    results = process_results(git_urls)
    save_to_excel(results, file_name)

if __name__ == "__main__":
    main()
